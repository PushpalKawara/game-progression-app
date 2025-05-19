import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from io import BytesIO
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
import tempfile

# Initialize Streamlit app
st.set_page_config(page_title="Game Analytics Tool", layout="wide")
st.title("🎮 Game Level Data Analyzer")

# ======================== DATA PROCESSING FUNCTIONS ========================
def clean_level(level):
    """Extract numeric value from LEVEL column"""
    if pd.isna(level):
        return 0
    return int(re.sub(r'\D', '', str(level)))

def process_files(start_df, complete_df):
    """Process and merge the two dataframes"""
    # Clean and sort data
    for df in [start_df, complete_df]:
        df['LEVEL'] = df['LEVEL'].apply(clean_level)
        df.sort_values('LEVEL', inplace=True)

    # Rename columns
    start_df = start_df.rename(columns={'USERS': 'Start Users'})
    complete_df = complete_df.rename(columns={'USERS': 'Complete Users'})

    # Merge data
    merge_cols = ['GAME_ID', 'DIFFICULTY', 'LEVEL']
    merged = pd.merge(start_df, complete_df, on=merge_cols, how='outer', suffixes=('_start', '_complete'))

    # Select required columns
    keep_cols = ['GAME_ID', 'DIFFICULTY', 'LEVEL', 'Start Users', 'Complete Users',
                 'PLAY_TIME_AVG', 'HINT_USED_SUM', 'SKIPPED_SUM', 'ATTEMPTS_SUM']
    merged = merged[keep_cols]

  # Calculate metrics
    merged['Game Play Drop'] = ((merged['Start Users'] - merged['Complete Users']) / merged['Start Users'].replace(0, np.nan)) * 100
    merged['Popup Drop'] = ((merged['Complete Users'] - merged['Start Users'].shift(-1)) / merged['Complete Users'].replace(0, np.nan)) * 100
    merged['Total Level Drop'] = ((merged['Start Users'] - merged['Start Users'].shift(-1)) / merged['Start Users'].replace(0, np.nan)) * 100
    merged['Retention %'] = (merged['Start Users'] / merged['Start Users'].max()) * 100
    # Fill NaN values
    merged.fillna({'Start Users': 0, 'Complete Users': 0}, inplace=True)
    return merged

# ======================== CHART GENERATION ========================
def create_charts(df, game_name):
    """Generate matplotlib charts"""
    charts = {}

    # Retention Chart
    fig1, ax1 = plt.subplots(figsize=(12, 4))
    ax1.plot(df['LEVEL'], df['Retention %'], color='#4CAF50')
    ax1.set_title(f"{game_name} - Retention %", fontsize=10)
    charts['retention'] = fig1

    # Total Level Drop Chart
    fig2, ax2 = plt.subplots(figsize=(12, 4))
    ax2.bar(df['LEVEL'], df['Total Level Drop'], color='#F44336')
    ax2.set_title(f"{game_name} - Total Level Drop", fontsize=10)
    charts['total_drop'] = fig2

    # Combined Drop Chart
    fig3, ax3 = plt.subplots(figsize=(12, 4))
    width = 0.35
    ax3.bar(df['LEVEL'] - width/2, df['Game Play Drop'], width, label='Game Play Drop')
    ax3.bar(df['LEVEL'] + width/2, df['Popup Drop'], width, label='Popup Drop')
    ax3.set_title(f"{game_name} - Drop Comparison", fontsize=10)
    ax3.legend()
    charts['combined_drop'] = fig3

    return charts

# ======================== CHART ADDITION TO EXCEL ========================
def add_charts_to_excel(worksheet, charts):
    """Add matplotlib charts to Excel worksheet as images"""
    img_positions = {
        'retention': 'M2',
        'total_drop': 'N32',
        'combined_drop': 'N65'
    }

    for chart_type in ['retention', 'total_drop', 'combined_drop']:
        # Save chart to bytes buffer
        img_data = BytesIO()
        charts[chart_type].savefig(img_data, format='png', dpi=150, bbox_inches='tight')
        img_data.seek(0)

        # Create image object
        img = OpenpyxlImage(img_data)

        # Add image to worksheet
        worksheet.add_image(img, img_positions[chart_type])

        # Close figure to prevent memory leaks
        plt.close(charts[chart_type])

# ======================== EXCEL GENERATION ========================
def generate_excel(processed_data):
    """Create Excel workbook with formatted sheets"""
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Create MAIN_TAB sheet
    main_sheet = wb.create_sheet("MAIN_TAB")
    main_headers = ["Index", "Sheet Name", "Game Play Drop Count", "Popup Drop Count",
                    "Total Level Drop Count", "LEVEL_Start", "Start Users",
                    "LEVEL_End", "USERS_END", "Link to Sheet"]
    main_sheet.append(main_headers)

    # Format main sheet headers
    for col in main_sheet[1]:
        col.font = Font(bold=True, color="FFFFFF")
        col.fill = PatternFill("solid", fgColor="4F81BD")

    # Process each game variant
    for idx, (game_id, df) in enumerate(processed_data.items(), start=1):
        sheet_name = f"{game_id}_{df['DIFFICULTY'].iloc[0]}"[:31]
        ws = wb.create_sheet(sheet_name)

        # Add backlink to MAIN_TAB
        ws['A1'] = '=HYPERLINK("#MAIN_TAB!A1", "Back to Main")'
        ws['A1'].font = Font(color="0000FF", underline="single")

        # Prepare data for sheet
        headers = ["Level", "Start Users", "Complete Users", "Game Play Drop",
                   "Popup Drop", "Total Level Drop", "Retention %",
                   "PLAY_TIME_AVG", "HINT_USED_SUM", "SKIPPED_SUM", "ATTEMPTS_SUM"]
        ws.append(headers)

        # Add data rows
        for _, row in df.iterrows():
            ws.append([
                row['LEVEL'], row['Start Users'], row['Complete Users'],
                row['Game Play Drop'], row['Popup Drop'], row['Total Level Drop'],
                row['Retention %'], row['PLAY_TIME_AVG'], row['HINT_USED_SUM'],
                row['SKIPPED_SUM'], row['ATTEMPTS_SUM']
            ])

        # Add charts
        charts = create_charts(df, sheet_name)
        add_charts_to_excel(ws, charts)

        # Formatting
        apply_sheet_formatting(ws)
        apply_conditional_formatting(ws, df.shape[0])

        # Update MAIN_TAB
        main_row = [
            idx, sheet_name,
            sum(df['Game Play Drop'] >= (df['Start Users'] * 0.03)),
            sum(df['Popup Drop'] >= (df['Start Users'] * 0.03)),
            sum(df['Total Level Drop'] >= (df['Start Users'] * 0.03)),
            df['LEVEL'].min(), df['Start Users'].max(),
            df['LEVEL'].max(), df['Complete Users'].iloc[-1],
            f'=HYPERLINK("#{sheet_name}!A1", "View")'
        ]
        main_sheet.append(main_row)

    # Format main sheet
    for col in range(1, len(main_headers)+1):
        main_sheet.column_dimensions[get_column_letter(col)].width = 18

    return wb

# ======================== REMAINING FUNCTIONS AND UI (UNCHANGED) ========================
# [Keep the apply_sheet_formatting, apply_conditional_formatting, and main() functions
# from the previous implementation unchanged]

def apply_sheet_formatting(sheet):
    """Apply consistent formatting to sheets"""
    # Freeze header row
    sheet.freeze_panes = 'A1'

    # Format headers
    for cell in sheet[1]:  # Data headers start at row 1
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")


    # Auto-fit columns
    for col in sheet.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        sheet.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

def apply_conditional_formatting(sheet, num_rows):
    """Apply color scale formatting to drop columns"""
    drop_columns = {'D', 'E', 'F'}  # Game Play Drop, Popup Drop, Total Level Drop

    red_scale = {
        '3': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
        '7': PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid'),
        '10': PatternFill(start_color='FF6666', end_color='FF6666', fill_type='solid')
    }

    for row in sheet.iter_rows(min_row=2, max_row=num_rows+1):
        for cell in row:
            if cell.column_letter in drop_columns and cell.value is not None:
                value = cell.value
                if value >= 10:
                    cell.fill = red_scale['10']
                elif value >= 7:
                    cell.fill = red_scale['7']
                elif value >= 3:
                    cell.fill = red_scale['3']
                cell.font = Font(color="FFFFFF")


     # Center alignment for all cells
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

# ======================== STREAMLIT UI ========================
def main():
    st.sidebar.header("Upload Files")
    start_file = st.sidebar.file_uploader("LEVEL_START.csv", type="csv")
    complete_file = st.sidebar.file_uploader("LEVEL_COMPLETE.csv", type="csv")

    if start_file and complete_file:
        with st.spinner("Processing data..."):
            try:
                # Read and process data
                start_df = pd.read_csv(start_file)
                complete_df = pd.read_csv(complete_file)
                merged = process_files(start_df, complete_df)

                # Group by game and difficulty
                processed_data = {}
                for (game_id, difficulty), group in merged.groupby(['GAME_ID', 'DIFFICULTY']):
                    processed_data[f"{game_id}"] = group

                # Generate Excel file
                wb = generate_excel(processed_data)

                # Save to bytes buffer
                with tempfile.NamedTemporaryFile(delete=False) as tmp:
                    wb.save(tmp.name)
                    with open(tmp.name, "rb") as f:
                        excel_bytes = f.read()

                # Download button
                st.success("Processing complete!")
                st.download_button(
                    label="📥 Download Consolidated Report",
                    data=excel_bytes,
                    file_name="Game_Analytics_Report.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

                # Show preview
                with st.expander("Preview Processed Data"):
                    st.dataframe(merged.head(20))

            except Exception as e:
                st.error(f"Error processing files: {str(e)}")

if __name__ == "__main__":
    main()
